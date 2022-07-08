import openpyxl
import os
from time import sleep
#创建一个工作簿
f = openpyxl.Workbook()

row=128           #行数
column=8         #列数
count=20359      #开始
max=65535        #最大
sheet_num=1
table_num=1
sheet='Sheet'+str(sheet_num)
f.create_sheet('Sheet'+str(sheet_num))
ws = f["Sheet"]
f.remove(ws)
begin=0
while count<=max+1:
    table = f[sheet]
    for column_tmp in range(column):
        for row_tmp in range(row-1):
            if count>max:
                f.save('btfs端口'+str(table_num)+'.xlsx')
                os._exit(0)
            value_tmp="sed -i 's/"+str(4003+row_tmp)+"/"+str(count)+"/' /root/.btfs"+str(row_tmp+2)+"/config"
            count+=1
            table.cell(row = row_tmp+1,column = column_tmp+1,value = value_tmp)
    sheet_num+=1
    sheet='Sheet'+str(sheet_num)
    f.save('Btfs端口'+str(table_num)+'.xlsx')
    table = f.create_sheet('Sheet'+str(sheet_num))
    if (sheet_num-1)%16==0:
        table_num+=1
        f = openpyxl.Workbook()
        sheet_num=1
        sheet='Sheet'+str(sheet_num)
        f.create_sheet('Sheet'+str(sheet_num))
        ws = f["Sheet"]
        f.remove(ws)
